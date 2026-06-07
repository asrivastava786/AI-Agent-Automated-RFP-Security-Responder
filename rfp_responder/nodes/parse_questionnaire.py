"""
nodes/parse_questionnaire.py – Ingest raw payload → structured QuestionItem list.

Supported input formats
───────────────────────
JSON  →  payload["format"] == "json"
         payload["questions"] = [{"question_text": "...", "category": "...", ...}, ...]

Excel →  payload["format"] == "excel"
         payload["file_content"] = base64-encoded .xlsx bytes
         payload["column_map"]   = {"question": "Question Text", "category": "Section"}

The node is idempotent: calling it twice with the same (thread_id, payload)
produces the same QuestionItem list because QuestionItem.create() derives
IDs deterministically from (thread_id, row_index).
"""

from __future__ import annotations

import base64
import io
import logging
from typing import Any

import structlog
from openpyxl import load_workbook

from rfp_responder.state import QuestionItem, QuestionStatus, RFPState, WorkflowStatus

logger = structlog.get_logger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Column map defaults for Excel ingestion
# ─────────────────────────────────────────────────────────────────────────────

_DEFAULT_COLUMN_MAP: dict[str, str] = {
    "question":    "Question",
    "category":    "Category",
    "control_id":  "Control ID",
    "context_hint": "Context / Notes",
}


# ─────────────────────────────────────────────────────────────────────────────
# Public node
# ─────────────────────────────────────────────────────────────────────────────

async def parse_questionnaire(state: RFPState) -> dict[str, Any]:
    """
    LangGraph node: parse_questionnaire

    Reads `state["raw_payload"]` and returns a partial state update containing:
      questions        – list[dict]  (QuestionItem.model_dump() per row)
      workflow_status  – WorkflowStatus.RETRIEVING  (signals next node)
      error_message    – populated only on fatal parse failure
    """
    log = logger.bind(
        thread_id=state["thread_id"],
        tenant_id=state["tenant_id"],
        questionnaire_id=state["questionnaire_id"],
    )

    payload: dict[str, Any] = state["raw_payload"]
    fmt: str = payload.get("format", "json").lower()

    try:
        if fmt == "json":
            raw_questions = _parse_json(payload)
        elif fmt == "excel":
            raw_questions = _parse_excel(payload)
        else:
            raise ValueError(f"Unsupported format '{fmt}'. Expected 'json' or 'excel'.")
    except Exception as exc:
        log.error("parse_questionnaire failed", error=str(exc))
        return {
            "workflow_status": WorkflowStatus.FAILED,
            "error_message": f"Questionnaire parse error: {exc}",
        }

    thread_id = state["thread_id"]
    items: list[dict[str, Any]] = []

    for idx, raw in enumerate(raw_questions):
        question_text = (raw.get("question_text") or raw.get("question") or "").strip()
        if not question_text:
            log.warning("Skipping blank question row", row_index=idx)
            continue

        item = QuestionItem.create(
            thread_id=thread_id,
            row_index=idx,
            question_text=question_text,
            category=raw.get("category", "General"),
            control_id=raw.get("control_id"),
            context_hint=raw.get("context_hint"),
        )
        items.append(item.model_dump())

    log.info("parse_questionnaire complete", question_count=len(items))

    return {
        "questions": items,
        "workflow_status": WorkflowStatus.RETRIEVING,
        "error_message": None,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Private parsers
# ─────────────────────────────────────────────────────────────────────────────

def _parse_json(payload: dict[str, Any]) -> list[dict[str, Any]]:
    """
    Extract a list of question dicts from a JSON payload.

    Accepts two shapes:
      {"questions": [...]}   – explicit key
      [...]                  – top-level array (legacy)
    """
    questions = payload.get("questions")
    if questions is None:
        raise KeyError("JSON payload missing required 'questions' key.")
    if not isinstance(questions, list):
        raise TypeError(f"'questions' must be a list, got {type(questions).__name__}.")
    return questions


def _parse_excel(payload: dict[str, Any]) -> list[dict[str, Any]]:
    """
    Decode a base64 .xlsx blob and extract rows using openpyxl.

    The column_map in the payload maps logical field names to the actual
    header strings found in the spreadsheet.
    """
    file_content_b64: str | None = payload.get("file_content")
    if not file_content_b64:
        raise ValueError("Excel payload missing 'file_content' (expected base64 string).")

    column_map: dict[str, str] = {
        **_DEFAULT_COLUMN_MAP,
        **payload.get("column_map", {}),
    }

    raw_bytes = base64.b64decode(file_content_b64)
    workbook  = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    sheet     = workbook.active

    if sheet is None:
        raise ValueError("Excel file has no active sheet.")

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel sheet is empty.")

    # First row is the header
    headers: list[str] = [str(h).strip() if h else "" for h in rows[0]]

    # Reverse-map: logical name → column index
    col_index: dict[str, int] = {}
    for logical, header_label in column_map.items():
        try:
            col_index[logical] = headers.index(header_label)
        except ValueError:
            pass   # Optional columns may be absent; required ones validated below

    if "question" not in col_index:
        raise ValueError(
            f"Could not find question column '{column_map['question']}' in headers {headers}."
        )

    def _get(row: tuple, key: str) -> str | None:
        idx = col_index.get(key)
        if idx is None or idx >= len(row):
            return None
        val = row[idx]
        return str(val).strip() if val is not None else None

    questions: list[dict[str, Any]] = []
    for row in rows[1:]:   # skip header row
        question_text = _get(row, "question")
        if not question_text:
            continue
        questions.append({
            "question_text": question_text,
            "category":      _get(row, "category") or "General",
            "control_id":    _get(row, "control_id"),
            "context_hint":  _get(row, "context_hint"),
        })

    workbook.close()
    return questions
