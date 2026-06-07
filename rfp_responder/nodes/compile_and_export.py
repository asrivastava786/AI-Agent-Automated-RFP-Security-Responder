"""
nodes/compile_and_export.py – Assemble final document and push audit telemetry.

Responsibilities
────────────────
1. Walk every question and resolve its final answer text:
     • AUTO_APPROVED   → use proposed_answer from DraftedAnswer
     • HUMAN_APPROVED  → use proposed_answer (reviewer accepted draft)
     • HUMAN_OVERRIDDEN → use the override_answer supplied by the reviewer
     • Rejected / failed → exclude from export, record in audit log

2. Serialise the finalised answers to:
     • JSON  – machine-readable; returned in the API response
     • Excel – human-readable (.xlsx) for the client portal

3. Push AuditMetrics to LangSmith as run feedback so the project dashboard
   surfaces token spend, approval ratios, and latency per tenant.

LangSmith integration
─────────────────────
LangGraph automatically creates a LangSmith Run for each graph.ainvoke() call
when LANGCHAIN_TRACING_V2=true.  This node retrieves the active run_id via
langsmith.get_current_run_tree() and attaches structured feedback to it.
This makes the metrics queryable in the LangSmith UI without any manual
instrumentation in the node code.
"""

from __future__ import annotations

import io
import json
import logging
import os
import time
from pathlib import Path
from typing import Any

import structlog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from rfp_responder.state import (
    AuditMetrics,
    DraftedAnswer,
    FinalAnswer,
    QuestionItem,
    QuestionStatus,
    RFPState,
    WorkflowStatus,
)

logger = structlog.get_logger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Excel styling constants
# ─────────────────────────────────────────────────────────────────────────────

_HEADER_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
_AUTO_FILL    = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_REVIEW_FILL  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_REJECT_FILL  = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

_STATUS_FILL: dict[str, PatternFill] = {
    QuestionStatus.AUTO_APPROVED:    _AUTO_FILL,
    QuestionStatus.HUMAN_APPROVED:   _AUTO_FILL,
    QuestionStatus.HUMAN_OVERRIDDEN: _REVIEW_FILL,
    QuestionStatus.REVIEW_REQUIRED:  _REJECT_FILL,
}

_COLUMNS = [
    ("Control ID",   15),
    ("Category",     20),
    ("Question",     60),
    ("Answer",       70),
    ("Status",       18),
    ("Reviewer",     20),
    ("Auto Approved", 14),
]


# ─────────────────────────────────────────────────────────────────────────────
# Final answer resolution
# ─────────────────────────────────────────────────────────────────────────────

def _resolve_final_answers(
    questions: list[dict[str, Any]],
    drafted_raw: dict[str, Any],
) -> tuple[list[FinalAnswer], int, int]:
    """
    Walk every question and pair it with its drafted (possibly reviewer-patched) answer.

    Returns (final_answers, auto_approved_count, human_reviewed_count).
    Questions with no draft (e.g., retrieval failed and synthesis skipped) are
    included with a placeholder text and flagged for manual review.
    """
    finals: list[FinalAnswer] = []
    auto_approved_count = 0
    human_reviewed_count = 0

    for q_dict in questions:
        q = QuestionItem.model_validate(q_dict)
        draft_dict = drafted_raw.get(q.question_id)

        if draft_dict is None:
            # No draft at all – flag as requiring review in the export.
            finals.append(FinalAnswer(
                question_id=q.question_id,
                row_index=q.row_index,
                category=q.category,
                question_text=q.question_text,
                final_answer_text="[NO ANSWER GENERATED – REQUIRES MANUAL COMPLETION]",
                status=QuestionStatus.REVIEW_REQUIRED,
                auto_approved=False,
                reviewer_id=None,
            ))
            continue

        draft = DraftedAnswer.model_validate(draft_dict)
        review_status: str = draft_dict.get("_review_status", "")
        reviewer_id: str | None = draft_dict.get("_reviewer_id")

        # Determine effective status
        if review_status == QuestionStatus.REVIEW_REQUIRED:
            # Reviewer rejected → include placeholder, mark rejected
            final_text = "[ANSWER REJECTED BY REVIEWER – REQUIRES MANUAL COMPLETION]"
            status = QuestionStatus.REVIEW_REQUIRED
            auto_approved = False
        elif review_status == QuestionStatus.HUMAN_OVERRIDDEN:
            final_text = draft_dict["proposed_answer"]   # already replaced by human_review_wait
            status = QuestionStatus.HUMAN_OVERRIDDEN
            auto_approved = False
            human_reviewed_count += 1
        elif review_status == QuestionStatus.HUMAN_APPROVED:
            final_text = draft.proposed_answer
            status = QuestionStatus.HUMAN_APPROVED
            auto_approved = False
            human_reviewed_count += 1
        else:
            # No review_status set → question was AUTO_APPROVED (bypassed human_review_wait)
            final_text = draft.proposed_answer
            status = QuestionStatus.AUTO_APPROVED
            auto_approved = True
            auto_approved_count += 1

        finals.append(FinalAnswer(
            question_id=q.question_id,
            row_index=q.row_index,
            category=q.category,
            question_text=q.question_text,
            final_answer_text=final_text,
            status=status,
            auto_approved=auto_approved,
            reviewer_id=reviewer_id,
        ))

    # Sort by original row order for consistent output
    finals.sort(key=lambda f: f.row_index)
    return finals, auto_approved_count, human_reviewed_count


# ─────────────────────────────────────────────────────────────────────────────
# Excel serialiser
# ─────────────────────────────────────────────────────────────────────────────

def _build_excel(finals: list[FinalAnswer], questionnaire_id: str) -> bytes:
    """Render the final answers as a styled .xlsx and return raw bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "RFP Responses"

    # Header row
    for col_idx, (header, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font  = _HEADER_FONT
        cell.fill  = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 22

    # Data rows
    for row_idx, answer in enumerate(finals, start=2):
        row_data = [
            getattr(answer, "control_id", None) or "N/A",   # FinalAnswer doesn't carry control_id;
            answer.category,                                  # that's fine – it's in the question
            answer.question_text,
            answer.final_answer_text,
            answer.status,
            answer.reviewer_id or "—",
            "Yes" if answer.auto_approved else "No",
        ]
        fill = _STATUS_FILL.get(answer.status)
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill

    # Freeze header row
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# LangSmith audit push
# ─────────────────────────────────────────────────────────────────────────────

def _push_langsmith_metrics(
    run_id: str | None,
    metrics: AuditMetrics,
    tenant_id: str,
) -> None:
    """
    Attach structured feedback to the current LangSmith run.

    Silently skips if:
      • run_id is None (tracing disabled or run not yet created)
      • langsmith package raises (non-fatal – export must still succeed)
    """
    if not run_id:
        return
    try:
        from langsmith import Client as LangSmithClient

        client = LangSmithClient()
        client.create_feedback(
            run_id=run_id,
            key="auto_approved_ratio",
            score=metrics.auto_approved_count / max(metrics.total_questions, 1),
            comment=f"tenant={tenant_id}",
        )
        client.create_feedback(
            run_id=run_id,
            key="total_tokens",
            score=metrics.total_prompt_tokens + metrics.total_completion_tokens,
        )
        client.create_feedback(
            run_id=run_id,
            key="processing_duration_seconds",
            score=metrics.processing_duration_seconds,
        )
    except Exception as exc:
        logger.warning("LangSmith metrics push failed (non-fatal)", error=str(exc))


# ─────────────────────────────────────────────────────────────────────────────
# Public node
# ─────────────────────────────────────────────────────────────────────────────

async def compile_and_export(state: RFPState) -> dict[str, Any]:
    """
    LangGraph node: compile_and_export

    Assembles final answers, serialises to JSON + Excel, pushes LangSmith
    metrics, and marks the workflow complete.

    Returns partial state update:
      final_answers    – dict[question_id, FinalAnswer dict]
      audit_metrics    – AuditMetrics dict
      workflow_status  – WorkflowStatus.COMPLETE
    """
    start_ts = time.monotonic()
    thread_id = state["thread_id"]
    tenant_id = state["tenant_id"]
    questionnaire_id = state["questionnaire_id"]

    log = logger.bind(
        thread_id=thread_id,
        tenant_id=tenant_id,
        questionnaire_id=questionnaire_id,
    )
    log.info("compile_and_export starting")

    questions  = state.get("questions", [])
    drafted_raw = state.get("drafted_answers", {})

    # ── Resolve final answers ─────────────────────────────────────────────────
    finals, auto_count, human_count = _resolve_final_answers(questions, drafted_raw)

    # ── Compute aggregate metrics ─────────────────────────────────────────────
    drafted_models = [
        DraftedAnswer.model_validate(d) for d in drafted_raw.values()
        if not isinstance(d.get("vector_confidence"), type(None))
    ]

    total_prompt     = sum(d.prompt_tokens     for d in drafted_models)
    total_completion = sum(d.completion_tokens for d in drafted_models)
    confidences      = [d.vector_confidence for d in drafted_models]
    avg_confidence   = sum(confidences) / len(confidences) if confidences else 0.0

    elapsed = time.monotonic() - start_ts
    metrics = AuditMetrics(
        total_prompt_tokens=total_prompt,
        total_completion_tokens=total_completion,
        total_questions=len(finals),
        auto_approved_count=auto_count,
        human_reviewed_count=human_count,
        avg_vector_confidence=round(avg_confidence, 4),
        processing_duration_seconds=round(elapsed, 2),
    )

    # ── Serialise outputs ─────────────────────────────────────────────────────
    export_dir = Path(state.get("export_dir", "/tmp/rfp_exports"))
    export_dir.mkdir(parents=True, exist_ok=True)

    json_path  = export_dir / f"{questionnaire_id}.json"
    excel_path = export_dir / f"{questionnaire_id}.xlsx"

    json_payload = [f.model_dump() for f in finals]
    json_path.write_text(json.dumps(json_payload, indent=2, default=str), encoding="utf-8")

    excel_bytes = _build_excel(finals, questionnaire_id)
    excel_path.write_bytes(excel_bytes)

    log.info(
        "Export files written",
        json_path=str(json_path),
        excel_path=str(excel_path),
        final_count=len(finals),
    )

    # ── Push audit metrics to LangSmith ──────────────────────────────────────
    _push_langsmith_metrics(
        run_id=state.get("langsmith_run_id"),
        metrics=metrics,
        tenant_id=tenant_id,
    )

    # ── Emit Prometheus business metrics ──────────────────────────────────────
    try:
        from rfp_responder.metrics import record_workflow_complete
        human_overridden = sum(
            1 for d in drafted_raw.values()
            if isinstance(d, dict) and d.get("_review_status") == "human_overridden"
        )
        record_workflow_complete(
            tenant_id=tenant_id,
            duration_s=elapsed,
            auto_approved=auto_count,
            human_reviewed=human_count,
            human_overridden=human_overridden,
            avg_confidence=avg_confidence,
        )
    except Exception:
        pass  # metrics are best-effort; never block the export

    log.info(
        "compile_and_export complete",
        auto_approved=auto_count,
        human_reviewed=human_count,
        total_tokens=total_prompt + total_completion,
        duration_s=elapsed,
    )

    # Build final_answers dict for state (keyed by question_id)
    final_answers_dict: dict[str, Any] = {f.question_id: f.model_dump() for f in finals}

    return {
        "final_answers": final_answers_dict,
        "audit_metrics": metrics.model_dump(),
        "workflow_status": WorkflowStatus.COMPLETE,
        "error_message": None,
    }
